# views.py
import openpyxl
from openpyxl import Workbook
from rest_framework.exceptions import ValidationError
from rest_framework.views import APIView
from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from django.db.models import F, Q
from datetime import date
import datetime
import io
from django.http import HttpResponse
from django.views import View
from rest_framework import status
from django.db.models import Count
from .utils import get_sql_field_type, apply_validation_rules
from django.db import connection
from decimal import Decimal, InvalidOperation
import re
import xlsxwriter
from django.utils import timezone
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import Product, Catalog, ProductField, ValidationRule, UploadedFile, Alert, ProductTableInfo, SubmissionInfo
from .serializers import ProductSerializer, CatalogSerializer, CatalogListSerializer, ProductFieldSerializer, ValidationRuleSerializer, UploadedFileSerializer
import logging
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags


class ProductListCreateView(generics.ListCreateAPIView):
    """
    List all products or create a new product
    """
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save()


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update or delete a product
    """
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]


class CatalogListCreateView(generics.ListCreateAPIView):
    """
    List all catalogs or create a new catalog
    """
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return CatalogListSerializer
        return CatalogSerializer

    def get_queryset(self):
        queryset = Catalog.objects.all()
        product_id = self.request.query_params.get('product_id', None)
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        return queryset
    
    def create_product_table_info(self, product):
        """
        Create the ProductTableInfo record for the given product after creating the catalog.
        """
        # Check if ProductTableInfo already exists for the product
        if not ProductTableInfo.objects.filter(product=product).exists():
            # Get the fields for the product (this assumes ProductField objects are related to Product)
            fields = ProductField.objects.filter(product=product)

            # Assuming the table name is based on the product ID (can be customized)
            table_name = f"product_{product.id}"

            # Create a ProductTableInfo record
            table_info = ProductTableInfo(
                product=product,
                table_name=table_name,
                fields=[field.name for field in fields]  # List of field names
            )
            table_info.save()

    def perform_create(self, serializer):
        # Save the catalog instance
        catalog = serializer.save()

        # Get the associated product
        product = catalog.product

        # Check if the table for this product already exists
        table_name = f"product_{product.id}"
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = %s)",
                [table_name]
            )
            exists = cursor.fetchone()[0]

        if not exists:
            # Retrieve fields for the product
            fields = ProductField.objects.filter(product=product)

            # Build the CREATE TABLE SQL query
            create_table_sql = f"CREATE TABLE {table_name} ("
            for field in fields:
                # Escape column name if it contains spaces or special characters
                column_name = f'"{field.name}"' if ' ' in field.name or any(c in field.name for c in [';', '"', "'"]) else field.name

                # Determine the SQL data type
                column_def = f"{column_name} {get_sql_field_type(field)}"

                # Apply length and decimal constraints
                if field.field_type == 'decimal':
                    max_decimal_places = None
                    try:
                        # Fetch validation rule for max decimal places
                        validation_rule = ValidationRule.objects.filter(product_field__product=product).first()
                        if validation_rule:
                            if validation_rule.has_max_decimal:
                                max_decimal_places = validation_rule.max_decimal_places
                    except ValidationRule.DoesNotExist:
                        pass  # No validation rule for max decimal places

                    if field.length:
                        column_def += f"({field.length}, {max_decimal_places if max_decimal_places else 0})"
                    else:
                        column_def += f"(10, {max_decimal_places if max_decimal_places else 0})"  # Default precision

                elif field.field_type == 'varchar' and field.length:
                    column_def += f"({field.length})"

                # Apply NOT NULL or NULL constraint
                column_def += " NOT NULL" if not field.is_null else " NULL"

                # Set primary key if applicable
                if field.is_primary_key:
                    column_def += " PRIMARY KEY"

                # Add the field definition to the table
                create_table_sql += f"{column_def}, "

            # Remove trailing comma and add closing parenthesis
            create_table_sql = create_table_sql.rstrip(", ") + ");"

            # Execute the SQL query to create the table
            with connection.cursor() as cursor:
                cursor.execute(create_table_sql)

            # Apply validation rules as constraints
            apply_validation_rules(fields, table_name)
            self.create_product_table_info(product)


    def create(self, request, *args, **kwargs):
        try:
            print("Request Data:", request.data)
            return super().create(request, *args, **kwargs)
        except Exception as e:
            print(str(e))
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

class CatalogDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update or delete a catalog
    """
    queryset = Catalog.objects.all()
    serializer_class = CatalogSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def perform_update(self, serializer):
        # Handle file upload and additional logic if needed
        serializer.save()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)

        try:
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


# Custom views for specific functionality
class ProductCatalogsView(generics.ListAPIView):
    """
    List all catalogs for a specific product
    """
    serializer_class = CatalogListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        product_id = self.kwargs['pk']
        return Catalog.objects.filter(product_id=product_id)


# Custom filter views
class CatalogFilterView(generics.ListAPIView):
    """
    Filter catalogs based on query parameters
    """
    serializer_class = CatalogListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Catalog.objects.all()

        # Add filtering options
        corporate = self.request.query_params.get('corporate', None)
        frequency = self.request.query_params.get('frequency', None)
        mandatory = self.request.query_params.get('mandatory', None)

        if corporate:
            queryset = queryset.filter(corporate__icontains=corporate)
        if frequency:
            queryset = queryset.filter(frequency=frequency)
        if mandatory:
            queryset = queryset.filter(mandatory=mandatory)

        return queryset

class FileUploadView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, catalog_id):
        try:
            catalog = Catalog.objects.get(id=catalog_id)
        except Catalog.DoesNotExist:
            return Response({"error": "Catalog not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check for file in the request
        if 'file' not in request.FILES:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        valid_mime_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        ]

        # Validate file type
        if file.content_type not in valid_mime_types:
            return Response(
                {"error": "Invalid file type. Please upload an Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )
        # Extract the domain from the catalog's product
        domain = getattr(catalog.product, 'domain', None)

        # Get the user from the request
        user = request.user if request.user.is_authenticated else None

        # Save the file record in the database
        uploaded_file = UploadedFile.objects.create(
            catalog=catalog,
            file=file,
            domain=domain,
            user=user
        )

        # Load the Excel file
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
        except Exception as e:
            return Response({"error": "Error reading Excel file: " + str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Parse and save data
        errors = []
        product = catalog.product  # Assuming a one-to-one relationship with Product

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Skip header row
            try:
                # Extract data from the Excel row
                field_name = row[0].value
                field_type = row[1].value
                length = row[2].value
                is_nullable = row[3].value
                is_primary_key = row[4].value
                is_unique = row[5].value
                is_picklist = row[6].value
                picklist_values = row[7].value
                has_min_max = row[8].value
                min_value = row[9].value
                max_value = row[10].value
                is_email_format = row[11].value
                is_phone_format = row[12].value
                has_max_decimal = row[13].value
                max_decimal_places = row[14].value
                has_date_format = row[15].value
                date_format = row[16].value
                has_max_days_of_age = row[17].value
                max_days_of_age = row[18].value
                custom_validation = row[19].value

                # Create ProductField
                product_field = ProductField.objects.create(
                    name=field_name,
                    field_type=field_type,
                    length=length,
                    is_null=bool(is_nullable),
                    is_primary_key=bool(is_primary_key),
                    product=product
                )
                # Create ValidationRule
                validation_rule = ValidationRule.objects.create(
                    product_field=product_field,
                    is_unique=bool(is_unique),
                    is_picklist=bool(is_picklist),
                    picklist_values=picklist_values,
                    has_min_max=bool(has_min_max),
                    min_value=min_value,
                    max_value=max_value,
                    is_email_format=bool(is_email_format),
                    is_phone_format=bool(is_phone_format),
                    has_max_decimal=bool(has_max_decimal),
                    max_decimal_places=max_decimal_places,
                    has_date_format=bool(has_date_format),
                    date_format=date_format,
                    has_max_days_of_age=bool(has_max_days_of_age),
                    max_days_of_age=max_days_of_age,
                    custom_validation=custom_validation
                )
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        if errors:
            return Response({"message": "File uploaded with some errors", "errors": errors}, status=status.HTTP_207_MULTI_STATUS)

        return Response(
            {"message": "File uploaded and data saved successfully"},
            status=status.HTTP_201_CREATED
        )
    

class RecentAndOverdueUploadsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Extract filters from the request body
        from_date_str = request.data.get('from')
        to_date_str = request.data.get('to')
        domain = request.data.get("domain")
        product = request.data.get("product")
        username = request.data.get("user")

        # Convert string dates to datetime objects
        from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d") if from_date_str else None
        to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d") if to_date_str else None

        # Validate date range: ensure from_date is earlier than to_date
        if from_date and to_date and from_date > to_date:
            raise ValidationError({"error": "The 'from' date must be earlier than the 'to' date."})

        # Build the filter conditions
        filters = Q()
        
        if from_date and to_date:
            filters &= Q(submission_time__date__range=[from_date, to_date])
        if domain:
            filters &= Q(domain=domain)
        if product:
            filters &= Q(product__schema_name=product)
        if username:
            filters &= Q(submitted_by__username=username)

        # Apply filters to the querysets
        new_submissions = SubmissionInfo.objects.filter(filters, submission_time__date=date.today()).count()
        total_submissions = SubmissionInfo.objects.filter(filters).count()
        delayed_submissions = SubmissionInfo.objects.filter(
            filters, submission_time__gt=F('catalog__deadline')
        ).count()

        # Fetch latest submissions within the filter scope
        latest_submissions = SubmissionInfo.objects.filter(filters).order_by('-submission_time')[:3]

        # Fetch overdue submissions within the filter scope
        overdue_submissions = SubmissionInfo.objects.filter(
            filters, submission_time__gt=F('catalog__deadline')
        ).order_by('-submission_time')[:3]

        # Fetch the alerts related to the user
        user_alerts = Alert.objects.filter(user=request.user).order_by('-created_at')

        # Serialize the data
        data = {
            "new_submissions": new_submissions,
            "total_submissions": total_submissions,
            "delayed_submissions": delayed_submissions,
            "last_submitted": [
                {
                    "date": submission.submission_time.date().strftime("%d/%m/%Y"),
                    "domain": submission.domain,
                    "user": submission.submitted_by.username if submission.submitted_by and submission.submitted_by.username else None
                } for submission in latest_submissions
            ],
            "delayed_records": [
                {
                    "date": submission.submission_time.date().strftime("%d/%m/%Y"),
                    "domain": submission.domain,
                    "status": submission.catalog.status if submission.catalog else None
                } for submission in overdue_submissions
            ],
            "alerts": [
                {
                    "message": alert.message,
                    "created_at": alert.created_at.date().strftime("%d/%m/%Y")
                } for alert in user_alerts
            ]
        }

        return Response(data, status=status.HTTP_200_OK)

class ProductFieldListCreateView(generics.ListCreateAPIView):
    """
    List all fields for a specific product or create a new field for that product.
    """
    serializer_class = ProductFieldSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        product_id = self.kwargs['product_id']
        return ProductField.objects.filter(product_id=product_id)

    def perform_create(self, serializer):
        # Retrieve the product ID from the URL and ensure it exists
        product_id = self.kwargs['product_id']
        try:
            product = Product.objects.get(id=product_id)
            # Pass the product instance when saving
            serializer.save(product=product)
        except Product.DoesNotExist:
            raise NotFound("Product not found.")


class ProductFieldDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update, or delete a specific field for a product.
    """
    serializer_class = ProductFieldSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        product_id = self.kwargs['product_id']
        field_id = self.kwargs['field_id']

        try:
            return ProductField.objects.get(id=field_id, product_id=product_id)
        except ProductField.DoesNotExist:
            raise NotFound("Product field not found.")


class ValidationRuleCreateView(generics.CreateAPIView):
    """
    Create a validation rule only if the associated product field exists.
    """
    serializer_class = ValidationRuleSerializer
    parser_classes = (JSONParser, FormParser, MultiPartParser)
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        product_id = self.kwargs['product_id']
        field_id = self.kwargs['field_id']

        # Verify if the ProductField exists
        try:
            product_field = ProductField.objects.get(id=field_id, product_id=product_id)
        except ProductField.DoesNotExist:
            raise NotFound("Product field not found.")

        # Create ValidationRule with the existing ProductField
        validation_rule_data = request.data
        validation_rule_data['product_field'] = product_field.id  # Associate product_field_id
        validation_rule_serializer = ValidationRuleSerializer(data=validation_rule_data)
        if validation_rule_serializer.is_valid():
            validation_rule = validation_rule_serializer.save()
            return Response(validation_rule_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(
                validation_rule_serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )


class ValidationRuleDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update, or delete a validation rule for a specific field of a product.
    """
    serializer_class = ValidationRuleSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        product_id = self.kwargs['product_id']
        field_id = self.kwargs['field_id']
        validation_rule_id = self.kwargs['validation_rule_id']
        try:
            validation_rule = ValidationRule.objects.get(id=validation_rule_id, product_field=field_id)
            if validation_rule.product_field.product.id != int(product_id):
                raise NotFound("Validation rule not found for this product field.")
            return validation_rule
        except ValidationRule.DoesNotExist:
            raise NotFound("Validation rule not found.")
        


class PendingCatalogsExportView(APIView):
    # Uncomment the next line if you want to require authentication
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Filter catalogs with 'Pending' status
            pending_catalogs = Catalog.objects.filter(status='Pending')
            
            # Create a new workbook and select the active worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Pending Catalogs'
            
            # Write headers
            headers = [
                'ID', 'Name', 'Corporate', 'Responsible User', 
                'Menu', 'Product', 'Mandatory', 'Frequency', 
                'Deadline', 'Submission Email', 'Created At'
            ]
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header)
            
            # Write catalog data
            for row, catalog in enumerate(pending_catalogs, start=2):
                worksheet.cell(row=row, column=1, value=catalog.id)
                worksheet.cell(row=row, column=2, value=catalog.name)
                worksheet.cell(row=row, column=3, value=catalog.corporate)
                worksheet.cell(row=row, column=4, value=str(catalog.responsible_user) if catalog.responsible_user else '')
                worksheet.cell(row=row, column=5, value=catalog.menu)
                worksheet.cell(row=row, column=6, value=str(catalog.product) if catalog.product else '')
                worksheet.cell(row=row, column=7, value=catalog.mandatory)
                worksheet.cell(row=row, column=8, value=catalog.frequency)
                worksheet.cell(row=row, column=9, value=catalog.deadline)
                worksheet.cell(row=row, column=10, value=catalog.submission_email)
                
                # Remove timezone information from created_at
                created_at_naive = catalog.created_at.replace(tzinfo=None) if catalog.created_at else None
                worksheet.cell(row=row, column=11, value=created_at_naive)
            
            # Create an in-memory bytes buffer to store the Excel file
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            
            # Create response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="pending_catalogs.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            
            return response
        
        except Exception as e:
            # Log the full error for debugging
            return Response({
                'error': 'Failed to generate export',
                'details': str(e)
            }, status=500)
        



class CatalogStatusCountView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
         # Use Django's annotations to count catalogs by status
        status_counts = Catalog.objects.values('status').annotate(count=Count('id'))
        
        # Convert QuerySet to a more readable dictionary
        status_summary = {
            item['status']: item['count'] for item in status_counts
        }
        
        # Ensure all possible statuses are represented, even if count is 0
        for status_choice, _ in Catalog.STATUS_CHOICES:
            if status_choice not in status_summary:
                status_summary[status_choice] = 0
        
        # Total catalogs
        total_catalogs = sum(status_summary.values())
        
        # Construct status_percentages with modified keys and exclude "Delayed"
        status_percentages = {}
        if total_catalogs > 0:
            status_percentages['Homologated'] = round((status_summary.get('Active', 0) / total_catalogs * 100), 2)
            status_percentages['Pending'] = round((status_summary.get('Pending', 0) / total_catalogs * 100), 2)
            status_percentages['Rejected'] = round((status_summary.get('Rejected', 0) / total_catalogs * 100), 2)
        
        # Get top three corporates by catalog count
        top_corporates = (
            Catalog.objects.values('corporate')
            .annotate(total_catalogs=Count('id'))
            .order_by('-total_catalogs')[:3]
        )

        # Summarize statuses for top corporates
        corporate_data = []
        for corp in top_corporates:
            corp_name = corp['corporate']
            corp_status_counts = (
                Catalog.objects.filter(corporate=corp_name)
                .values('status')
                .annotate(count=Count('id'))
            )

            # Convert QuerySet to a dictionary
            corp_status_summary = {item['status']: item['count'] for item in corp_status_counts}

            # Ensure all relevant statuses are included
            corp_status_summary = {
                'Homologated': corp_status_summary.get('Active', 0),
                'Pending': corp_status_summary.get('Pending', 0),
                'Rejected': corp_status_summary.get('Rejected', 0),
            }

            corporate_data.append({
                'corporate': corp_name,
                'total_catalogs': corp['total_catalogs'],
                'status_summary': corp_status_summary,
            })


        # Prepare response
        response_data = {
            'status_counts': status_summary,
            'total_catalogs': total_catalogs,
            'status_percentages': status_percentages,
            'top_corporates': corporate_data,
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
    


class UploadedFileListView(generics.ListAPIView):
    serializer_class = UploadedFileSerializer
    
    def get_queryset(self):
        # Extract catalog_id from the URL
        catalog_id = self.kwargs['catalog_id']
        
        # Filter UploadedFile records based on the provided catalog_id
        return UploadedFile.objects.filter(catalog_id=catalog_id)


class DynamicTableDataListView(APIView):
    """
    API to retrieve all rows from a dynamic table for a specific product.
    URL: /product/{product_id}/product-data/
    """
    def get(self, request, product_id):
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get the associated dynamic table (only one table per product)
        product_table = ProductTableInfo.objects.filter(product=product).first()

        if not product_table:
            return Response({'error': 'No dynamic table found for this product'}, status=status.HTTP_404_NOT_FOUND)
        

        # Fetch all product fields for the product
        product_fields = ProductField.objects.filter(product=product)

        # Gather comprehensive field details
        field_details = {}
        for field in product_fields:
            validation_rule = ValidationRule.objects.filter(product_field=field).first()
            field_info = {
                'type': field.field_type,
                'length': field.length,
                'is_null': field.is_null,
                'is_primary_key': field.is_primary_key,
                'is_unique': validation_rule.is_unique if validation_rule else None,
                'is_picklist': validation_rule.is_picklist if validation_rule else None,
                'picklist_values': validation_rule.picklist_values.split(',') if validation_rule and validation_rule.picklist_values else [],
                'has_min_max': validation_rule.has_min_max if validation_rule else None,
                'min_value': validation_rule.min_value if validation_rule else None,
                'max_value': validation_rule.max_value if validation_rule else None,
                'is_email_format': validation_rule.is_email_format if validation_rule else None,
                'is_phone_format': validation_rule.is_phone_format if validation_rule else None,
                'has_max_decimal': validation_rule.has_max_decimal if validation_rule else None,
                'max_decimal_places': validation_rule.max_decimal_places if validation_rule else None,
                'has_date_format': validation_rule.has_date_format if validation_rule else None,
                'date_format': validation_rule.date_format if validation_rule else None,
                'has_max_days_of_age': validation_rule.has_max_days_of_age if validation_rule else None,
                'max_days_of_age': validation_rule.max_days_of_age if validation_rule else None,
                'custom_validation': validation_rule.custom_validation if validation_rule else None,
            }
            field_details[field.name] = field_info

        # Fetch all rows from the dynamic table for the given product
        data = {}
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {product_table.table_name}")
            columns = [col[0] for col in cursor.description]
            column_types = [col[1] for col in cursor.description]  # Extract column types

            column_type_mapping = {
                'INTEGER': 'Integer',
                'VARCHAR': 'String',
                'TEXT': 'Text',
                'BOOLEAN': 'Boolean',
                'DATE': 'Date',
                'TIMESTAMP': 'Timestamp',
                'DECIMAL': 'Decimal',
            }

            field_types = [column_type_mapping.get(str(col_type), str(col_type)) for col_type in column_types]


            # Include comprehensive table information in the response
            data[product_table.table_name] = {
                'fields': product_table.fields,  # Original fields
                'field_types': field_types,  # SQL field types
                'field_details': field_details
            }

        return Response(data, status=status.HTTP_200_OK)
    

class DynamicTableDataSaveView(APIView):
    """
    API to save data to a dynamic table for a specific product
    URL: /product/{product_id}/save-data/
    """
    def post(self, request, product_id):
        try:
            # Verify product exists
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get the associated dynamic table
        product_table = ProductTableInfo.objects.filter(product=product).first()

        if not product_table:
            return Response({'error': 'No dynamic table found for this product'}, status=status.HTTP_404_NOT_FOUND)

        # Get the product fields to map data types
        product_fields = ProductField.objects.filter(product=product)
        
        # Get the data from the request
        data = request.data
        print("Received Data:", data)

        # Initialize the list for converted values
        converted_values = []

        # Prepare the SQL insert statement
        columns = ', '.join([f'"{field}"' for field in data.keys()])  # Wrap column names in double quotes
        placeholders = ', '.join(['%s'] * len(data))

        # Convert the values based on the field type in ProductField
        for field_name, value in data.items():
            field = product_fields.filter(name=field_name).first()  # Get the field metadata
            if field:
                field_type = field.field_type

                if field_type == 'varchar' or field_type == 'text':
                    # For varchar or text, keep the value as a string
                    converted_values.append(str(value))

                elif field_type == 'int':
                    # For integer fields, convert to integer
                    try:
                        converted_values.append(int(value))
                    except ValueError:
                        converted_values.append(None)  # Or handle as per your need

                elif field_type == 'float':
                    # For float fields, convert to float
                    try:
                        converted_values.append(float(value))
                    except ValueError:
                        converted_values.append(None)  # Or handle as per your need

                elif field_type == 'boolean':
                    # For boolean fields, convert to True/False
                    converted_values.append(value.lower() in ['true', '1', 't', 'y', 'yes'])

                elif field_type == 'date':
                    # For date fields, convert to date (YYYY-MM-DD format)
                    try:
                        converted_values.append(value)  # You can use datetime.strptime(value, "%Y-%m-%d") if you need parsing
                    except ValueError:
                        converted_values.append(None)  # Or handle as per your need

                elif field_type == 'datetime':
                    # For datetime fields, you can parse to datetime (if required)
                    try:
                        from datetime import datetime
                        converted_values.append(datetime.strptime(value, "%Y-%m-%d %H:%M:%S"))
                    except ValueError:
                        converted_values.append(None)

                elif field_type == 'decimal':
                    # For decimal fields, convert to Decimal
                    try:
                        from decimal import Decimal
                        converted_values.append(Decimal(value))
                    except ValueError:
                        converted_values.append(None)

                else:
                    converted_values.append(value)  # Default case

            else:
                # If no field metadata is found for the field name, append the value as is
                converted_values.append(value)

        try:
            with connection.cursor() as cursor:
                # Construct and execute the dynamic insert SQL
                insert_query = f"INSERT INTO {product_table.table_name} ({columns}) VALUES ({placeholders})"
                print("SQL Query:", insert_query)
                print("Values to Insert:", converted_values)
                cursor.execute(insert_query, converted_values)

            return Response({'message': 'Data saved successfully'}, status=status.HTTP_201_CREATED)
        except Exception as e:
            print(str(e))
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    

class DynamicTableDataDetailView(APIView):
    """
    API to retrieve a specific row from a dynamic table for a specific product.
    URL: /product/{product_id}/product-data/{product_data_id}/
    """
    def get(self, request, product_id, product_data_id):
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get the associated dynamic table
        try:
            product_table = ProductTableInfo.objects.get(product=product, id=product_data_id)
        except ProductTableInfo.DoesNotExist:
            return Response({'error': 'Dynamic table not found'}, status=status.HTTP_404_NOT_FOUND)

        # Fetch all rows from the specific dynamic table
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT * FROM {product_table.table_name}")
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            data = [dict(zip(columns, row)) for row in rows]

        return Response(data, status=status.HTTP_200_OK)
    




class DynamicTableExcelValidationView(APIView):
    """
    API to validate Excel data against dynamic table schema
    URL: /product/{product_id}/validate-excel/
    """
    def post(self, request, product_id):
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get the associated dynamic table
        product_table = ProductTableInfo.objects.filter(product=product).first()
        if not product_table:
            return Response({'error': 'No dynamic table found for this product'}, status=status.HTTP_404_NOT_FOUND)

        # Get product fields
        product_fields = ProductField.objects.filter(product=product)
        
        # Get data and headers from request
        excel_data = request.data.get('data', [])
        print(excel_data)
        headers = request.data.get('headers', [])
        headers = [re.sub(r'\(.*?\)|\[.*?\]', '', header).strip() for header in headers]

        # Validation errors list
        validation_errors = []

        # Validate headers match expected fields
        expected_fields = [field.name for field in product_fields]
        for i, header in enumerate(headers):
            if header not in expected_fields:
                validation_errors.append({
                    'row': 0,
                    'field': header,
                    'error': f'Unexpected column header: {header}'
                })

        # Validate each row
        for row_index, row_data in enumerate(excel_data, start=1):
            # Create a dictionary mapping headers to values
            row_dict = dict(zip(headers, row_data))

            # Validate each field
            for field in product_fields:
                value = row_dict.get(field.name, '')
                
                # Get validation rules
                validation_rule = ValidationRule.objects.filter(product_field=field).first()
                
                # Specific validations based on field type and rules
                try:
                    if field.field_type == 'varchar' or field.field_type == 'text':
                        # Text length validation
                        if validation_rule and field.length and len(str(value)) > field.length:
                            validation_errors.append({
                                'row': row_index,
                                'field': field.name,
                                'error': f'Exceeds maximum length of {field.length}'
                            })
                        
                        # Email format validation
                        if validation_rule and validation_rule.is_email_format:
                            if not re.match(r'\S+@\S+\.\S+', str(value)):
                                validation_errors.append({
                                    'row': row_index,
                                    'field': field.name,
                                    'error': 'Invalid email format'
                                })

                    elif field.field_type == 'int':
                        # Integer validation
                        try:
                            int_value = int(value)
                            
                            # Min-max validation
                            if validation_rule:
                                if validation_rule.min_value is not None and int_value < validation_rule.min_value:
                                    validation_errors.append({
                                        'row': row_index,
                                        'field': field.name,
                                        'error': f'Value must be >= {validation_rule.min_value}'
                                    })
                                if validation_rule.max_value is not None and int_value > validation_rule.max_value:
                                    validation_errors.append({
                                        'row': row_index,
                                        'field': field.name,
                                        'error': f'Value must be <= {validation_rule.max_value}'
                                    })
                        except ValueError:
                            validation_errors.append({
                                'row': row_index,
                                'field': field.name,
                                'error': 'Must be an integer'
                            })

                    elif field.field_type == 'decimal':
                        # Decimal validation
                        try:
                            decimal_value = Decimal(str(value))
                            
                            # Decimal places validation
                            if validation_rule and validation_rule.max_decimal_places is not None:
                                decimal_str = str(decimal_value)
                                if len(decimal_str.split('.')[-1]) > validation_rule.max_decimal_places:
                                    validation_errors.append({
                                        'row': row_index,
                                        'field': field.name,
                                        'error': f'Maximum {validation_rule.max_decimal_places} decimal places allowed'
                                    })
                            
                            # Min-max validation
                            if validation_rule:
                                if validation_rule.min_value is not None and decimal_value < Decimal(str(validation_rule.min_value)):
                                    validation_errors.append({
                                        'row': row_index,
                                        'field': field.name,
                                        'error': f'Value must be >= {validation_rule.min_value}'
                                    })
                                if validation_rule.max_value is not None and decimal_value > Decimal(str(validation_rule.max_value)):
                                    validation_errors.append({
                                        'row': row_index,
                                        'field': field.name,
                                        'error': f'Value must be <= {validation_rule.max_value}'
                                    })
                        except InvalidOperation:
                            validation_errors.append({
                                'row': row_index,
                                'field': field.name,
                                'error': 'Must be a valid decimal number'
                            })

                    elif field.field_type == 'date':
                        date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']
                        date_valid = False
                        for date_format in date_formats:
                            try:
                                datetime.datetime.strptime(str(value), date_format)
                                date_valid = True
                                break 
                            except ValueError:
                                continue  # Try the next format

                        if not date_valid:
                            validation_errors.append({
                                'row': row_index,
                                'field': field.name,
                                'error': f'Invalid data format. Date format must be: {validation_rule.date_format}'
                            })


                    elif field.field_type == 'boolean':
                        # Boolean validation
                        if str(value).lower() not in ['true', 'false', '0', '1']:
                            validation_errors.append({
                                'row': row_index,
                                'field': field.name,
                                'error': 'Must be true/false or 0/1'
                            })

                except Exception as e:
                    validation_errors.append({
                        'row': row_index,
                        'field': field.name,
                        'error': str(e)
                    })

        # If errors found, return them
        if validation_errors:
            # Send email about validation errors
            email_sent = EmailService.send_validation_error_email(
                user=request.user, 
                product=product, 
                validation_errors=validation_errors
            )

            # Create alert for the user
            alert_created = AlertService.create_alert(
                user=request.user,
                message=f'Excel validation failed for {product.schema_name}. {len(validation_errors)} errors detected.'
            )
            return Response({
                'errors': validation_errors, 
                'email_sent': email_sent,
                'alert_created': bool(alert_created)
            }, status=status.HTTP_400_BAD_REQUEST)

        # If no errors, return success
        return Response({'message': 'Validation successful'}, status=status.HTTP_200_OK)
    

class DynamicTableExcelSaveView(APIView):
    """
    API to save Excel data to a dynamic table
    URL: /product/{product_id}/save-excel-data/
    """
    def post(self, request, product_id):
        try:
            # Verify product exists
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get the associated dynamic table
        product_table = ProductTableInfo.objects.filter(product=product).first()
        if not product_table:
            return Response({'error': 'No dynamic table found for this product'}, status=status.HTTP_404_NOT_FOUND)

        # Get product fields
        product_fields = ProductField.objects.filter(product=product)
        field_names = [field.name for field in product_fields]
        
        # Get data from request
        excel_data = request.data.get('data', [])
        print(excel_data)
        if not excel_data:
            return Response({'error': 'No data provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Prepare bulk insert data
        bulk_insert_data = []
        for row_data in excel_data:
            # Convert data based on field types
            if isinstance(row_data, list):  # If row_data is a list, map it to the field names
                row_dict = dict(zip(field_names, row_data))
            elif isinstance(row_data, dict):  # If row_data is already a dictionary, use it as-is
                row_dict = row_data
            else:
                continue  # Skip invalid rows
            converted_row = []
            for field in product_fields:
                value = row_dict.get(field.name)
                
                # Type conversion logic
                if field.field_type == 'int':
                    try:
                        converted_row.append(int(value))
                    except (ValueError, TypeError):
                        converted_row.append(None)

                elif field.field_type == 'float':
                    try:
                        converted_row.append(float(value))
                    except (ValueError, TypeError):
                        converted_row.append(None)

                elif field.field_type == 'decimal':
                    try:
                        from decimal import Decimal
                        converted_row.append(Decimal(str(value)))
                    except (ValueError, TypeError):
                        converted_row.append(None)

                elif field.field_type == 'boolean':
                    # Handle various boolean representations
                    if str(value).lower() in ['true', '1', 't', 'y', 'yes']:
                        converted_row.append(True)
                    elif str(value).lower() in ['false', '0', 'f', 'n', 'no']:
                        converted_row.append(False)
                    else:
                        converted_row.append(None)

                elif field.field_type == 'date':
                    try:
                        # Try multiple date formats
                        from datetime import datetime
                        formats = [
                            "%m/%d/%Y",   # MM/DD/YYYY (your current input format)
                            "%Y-%m-%d",   # YYYY-MM-DD
                            "%d/%m/%Y"    # DD/MM/YYYY
                        ]
                        
                        for fmt in formats:
                            try:
                                converted_date = datetime.strptime(str(value), fmt).date()
                                converted_row.append(converted_date)
                                break
                            except ValueError:
                                continue
                        else:
                            converted_row.append(None)
                    except (ValueError, TypeError):
                        converted_row.append(None)

                elif field.field_type == 'datetime':
                    try:
                        from datetime import datetime
                        # Try multiple common datetime formats
                        formats = [
                            "%Y-%m-%d %H:%M:%S",
                            "%Y-%m-%d %H:%M",
                            "%Y-%m-%d",
                            "%d/%m/%Y %H:%M:%S",
                            "%d/%m/%Y"
                        ]
                        
                        for fmt in formats:
                            try:
                                converted_row.append(datetime.strptime(str(value), fmt))
                                break
                            except ValueError:
                                continue
                        else:
                            converted_row.append(None)
                    except (ValueError, TypeError):
                        converted_row.append(None)

                else:
                    # For varchar, text, and other types, convert to string
                    converted_row.append(str(value) if value is not None else None)
            bulk_insert_data.append(converted_row)
        try:
            with connection.cursor() as cursor:
                # Prepare bulk insert SQL
                columns = ', '.join([f'"{field.name}"' for field in product_fields])
                placeholders = ', '.join(['%s'] * len(product_fields))
                
                # Construct and execute the bulk insert
                insert_query = f"INSERT INTO {product_table.table_name} ({columns}) VALUES ({placeholders})"
                print(insert_query)
                print(bulk_insert_data)
                
                # Execute bulk insert
                cursor.executemany(insert_query, bulk_insert_data)

            # Save submission info
            catalog=product.catalogs.first()
            submission = SubmissionInfo.objects.create(
                product=product,
                submitted_by=request.user,
                submitted_data=excel_data,
                submission_time=timezone.now(),
                catalog=catalog,
                domain=product.domain
            )

            return Response({
                'message': 'Excel data saved successfully',
                'rows_inserted': len(bulk_insert_data)
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            # Log the full error for server-side debugging
            error_message = str(e)
            print(f"Excel upload error: {error_message}")
            
            # Send email about save errors
            email_sent = EmailService.send_save_error_email(
                user=request.user, 
                product=product, 
                save_errors=error_message
            )

            # Create alert for the user
            alert_created = AlertService.create_alert(
                user=request.user,
                message=f'Excel data save failed for {product.name}. Error: {error_message}'
            )
            
            return Response({
                'error': 'Failed to save Excel data',
                'details': error_message,
                'email_sent': email_sent,
                'alert_created': bool(alert_created)
            }, status=status.HTTP_400_BAD_REQUEST)
        



class ExcelTemplateGenerationView(APIView):
    """
    API to generate an Excel template for a specific product's dynamic table
    URL: /product/{product_id}/excel-template/
    """
    def get(self, request, product_id):
        try:
            # Verify product exists
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get product fields
        product_fields = ProductField.objects.filter(product=product)

        # Create an in-memory Excel file
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Data')

        # Add headers with field details
        header_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#DCE6F1',  # Light blue background
            'border': 1
        })

        # Write headers with additional validation info
        headers = []
        for field in product_fields:
            # Get validation rules
            validation_rule = ValidationRule.objects.filter(product_field=field).first()
            
            # Construct header with validation details
            header_text = field.name
            
            # Add type information
            header_text += f" ({field.field_type})"
            
            # Add validation details
            additional_info = []
            if validation_rule:
                if validation_rule.min_value is not None:
                    additional_info.append(f"Min: {validation_rule.min_value}")
                if validation_rule.max_value is not None:
                    additional_info.append(f"Max: {validation_rule.max_value}")
                if not field.is_null:
                    additional_info.append("Required")
                if field.length:
                    if field.field_type == "text" or field.field_type == "varchar":
                        additional_info.append(f"Length: {field.length}")
                if validation_rule.is_picklist:
                    additional_info.append(f"Picklist: {validation_rule.picklist_values}")
                if field.field_type == "decimal" and validation_rule.has_max_decimal:
                    additional_info.append(f"Decimal Places: {validation_rule.max_decimal_places}")
                if field.field_type == 'date' and validation_rule.date_format:
                    date_format = validation_rule.date_format
                    if date_format == 'MM/DD/YYYY':
                        additional_info.append("Date Format: MM/DD/YYYY")
                    elif date_format == 'DD/MM/YYYY':
                        additional_info.append("Date Format: DD/MM/YYYY")
                    elif date_format == 'YYYY-MM-DD':
                        additional_info.append("Date Format: YYYY-MM-DD")
                    elif date_format == 'MM-DD-YYYY':
                        additional_info.append("Date Format: MM-DD-YYYY")
                    elif date_format == 'DD-MM-YYYY':
                        additional_info.append("Date Format: DD-MM-YYYY")
                    else:
                        additional_info.append("Date Format: Default (YYYY-MM-DD)")
            
            if additional_info:
                header_text += f" [{', '.join(additional_info)}]"
            
            headers.append(header_text)

        # Write headers
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Add data validation and formatting based on field types
        for col, field in enumerate(product_fields):
            validation_rule = ValidationRule.objects.filter(product_field=field).first()
            
            # Set column width
            worksheet.set_column(col, col, 20)

            # Add specific validations
            if field.field_type == 'int':
                # Integer validation
                if validation_rule:
                    if validation_rule.min_value is not None and validation_rule.max_value is not None:
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'integer',
                            'criteria': 'between',
                            'minimum': validation_rule.min_value,
                            'maximum': validation_rule.max_value
                        })
            
            elif field.field_type == 'decimal':
                # Decimal validation
                if validation_rule and validation_rule.max_decimal_places is not None:
                    worksheet.data_validation(1, col, 1048576, col, {
                        'validate': 'decimal',
                        'criteria': 'between',
                        'minimum': validation_rule.min_value or 0,
                        'maximum': validation_rule.max_value or float('inf')
                    })
            
            elif field.field_type == 'boolean':
                # Boolean dropdown
                worksheet.data_validation(1, col, 1048576, col, {
                    'validate': 'list',
                    'source': ['true', 'false']
                })
            
            elif field.field_type == 'date':
                # Date format validation
                if validation_rule and validation_rule.date_format:
                    date_format = validation_rule.date_format
                    
                    # Apply date format validation based on the date_format field
                    if date_format == 'MM/DD/YYYY':
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '01/01/2000',
                            'maximum': '12/31/2099'
                        })
                    elif date_format == 'DD/MM/YYYY':
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '01/01/2000',
                            'maximum': '31/12/2099'
                        })
                    elif date_format == 'YYYY-MM-DD':
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '2000-01-01',
                            'maximum': '2099-12-31'
                        })
                    elif date_format == 'MM-DD-YYYY':
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '01-01-2000',
                            'maximum': '12-31-2099'
                        })
                    elif date_format == 'DD-MM-YYYY':
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '01-01-2000',
                            'maximum': '31-12-2099'
                        })
                    else:
                        # Default validation if the date_format is not recognized
                        worksheet.data_validation(1, col, 1048576, col, {
                            'validate': 'date',
                            'criteria': 'between',
                            'minimum': '2000-01-01',
                            'maximum': '2099-12-31'
                        })
            
            elif validation_rule and validation_rule.is_picklist:
                # Picklist validation
                picklist_values = validation_rule.picklist_values.split(',')
                worksheet.data_validation(1, col, 1048576, col, {
                    'validate': 'list',
                    'source': picklist_values
                })

        # Close the workbook
        workbook.close()
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            output.read(), 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=product_{product_id}_template.xlsx'
        
        return response
    

class SubmissionListView(generics.ListAPIView):
    """
    API to list all submissions for a specific product.
    URL: /product/{product_id}/submissions/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, product_id):
        submissions = SubmissionInfo.objects.filter(product_id=product_id).select_related('submitted_by')
        data = [
            {
                'id': submission.id,
                'submitted_by': submission.submitted_by.username,
                'submission_time': submission.submission_time.strftime('%Y-%m-%d %H:%M:%S'),
                'submitted_data': submission.submitted_data,
            }
            for submission in submissions
        ]
        return Response(data, status=status.HTTP_200_OK)
    


class DeleteCatalogView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, catalog_id):
        if not request.user.is_staff:
            return Response(
                {"message": "Unauthorized: Admin access required."},
                status=status.HTTP_403_FORBIDDEN,
            )

        catalog = get_object_or_404(Catalog, id=catalog_id)

        catalog.delete()
        return Response(
            {"message": f"Catalog '{catalog.name}' deleted successfully."},
            status=status.HTTP_200_OK,
        )
    


class EmailService:
    @staticmethod
    def send_validation_error_email(user, product, validation_errors):
        """
        Send email to user about validation errors
        """
        try:
            # Prepare email context
            context = {
                'username': user.get_full_name() or user.username,
                'product_name': product.schema_name,
                'validation_errors': validation_errors
            }

            # Render HTML email template
            html_message = render_to_string('emails/validation_errors.html', context)
            plain_message = strip_tags(html_message)
            print(user.email)
            print(html_message)
            print(plain_message)

            # Send email
            send_mail(
                subject=f'Excel Validation Errors for {product.schema_name}',
                message=plain_message,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            return True
        except Exception as e:
            logging.error(f"Email sending failed: {str(e)}")
            return False

    @staticmethod
    def send_save_error_email(user, product, save_errors):
        """
        Send email to user about data save errors
        """
        try:
            # Prepare email context
            context = {
                'username': user.get_full_name() or user.username,
                'product_name': product.schema_name,
                'save_errors': save_errors
            }

            # Render HTML email template
            html_message = render_to_string('emails/save_errors.html', context)
            plain_message = strip_tags(html_message)
            print(user.email)
            print(html_message)
            print(plain_message)

            # Send email
            send_mail(
                subject=f'Data Save Errors for {product.schema_name}',
                message=plain_message,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            return True
        except Exception as e:
            logging.error(f"Email sending failed: {str(e)}")
            return False

class AlertService:
    @staticmethod
    def create_alert(user, message):
        """
        Create an alert for the user
        """
        try:
            alert = Alert.objects.create(
                user=user,
                message=message
            )
            return alert
        except Exception as e:
            logging.error(f"Alert creation failed: {str(e)}")
            return None
